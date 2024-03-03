import os, io, django, tempfile
from django.http import HttpResponse
from django.db.models import Q
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from .models import Equipo, Usuario, Software, Driver, Revision
from .forms import EquipoForm, UsuarioForm, SoftwareForm, DriverForm, RevisionForm
from fpdf import FPDF
from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook
from imagekitio import ImageKit
from imagekitio.models.UploadFileRequestOptions import UploadFileRequestOptions


@login_required
def home(request):
    search = request.GET.get("q")
    page_num = request.GET.get("page", 1)
    cont = err = 0

    if search:
        equipos = Equipo.objects.filter(
            Q(serial__icontains=search)
            | Q(usuario__ciudad__icontains=search)
            | Q(usuario__nombre__icontains=search)
        )
    else:
        equipos = Equipo.objects.all()
    page = Paginator(object_list=equipos, per_page=5).get_page(page_num)
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[4]:
                USUARIO = row[4].replace("_", " ").title()
                CIUDAD = row[7].replace("_", " ").title()
                TIPO_EQUIPO = row[8]
                MARCA = row[9]
                MODELO = row[10]
                SERIAL = row[12]
                ACTIVO = row[13]
                try:
                    equipo = Equipo.objects.create(
                        tipo=TIPO_EQUIPO, marca=MARCA, modelo=MODELO, serial=SERIAL
                    )
                    cont += 1
                    Usuario.objects.create(
                        nombre=USUARIO, ciudad=CIUDAD, activo=ACTIVO, equipo=equipo
                    )
                except django.db.IntegrityError:
                    print(
                        f"El equipo con serial {SERIAL} ya existe en la base de datos."
                    )
                    err += 1
        wb.close()
        # return render(request, "mantenimiento/home.html", context={"page": page})
    return render(
        request=request,
        template_name="mantenimiento/home.html",
        context={"page": page, "cont": cont, "err": err},
    )


@login_required
def usuario(request, pk):
    equipo = get_object_or_404(Equipo, pk=pk)
    usuario, created = Usuario.objects.get_or_create(equipo=equipo)
    if request.method == "POST":
        form = UsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            usuario.save()
            dibujarusuario(usuario)
            return redirect("equipo", pk=pk)
    else:
        form = UsuarioForm(instance=usuario)
    return render(request, "mantenimiento/usuario.html", {"form": form})


@login_required
def equipo(request, pk):
    equipo = get_object_or_404(Equipo, pk=pk)
    context = {}
    error_img = "Faltan las siguientes fotos: "
    if request.method == "POST":
        form = EquipoForm(request.POST, request.FILES, instance=equipo)
        if form.is_valid():
            equipo = form.save(commit=False)
            # Subir imagenes
            front = request.FILES.get("input_in_front")
            back = request.FILES.get("input_in_back")
            right1 = request.FILES.get("input_in_right1")
            right2 = request.FILES.get("input_in_right2")
            left1 = request.FILES.get("input_in_left1")
            left2 = request.FILES.get("input_in_left2")
            if front or back or right1 or right2 or left1 or left2:
                imagekit = ImageKit(
                    private_key="private_OHsLbCwevXZ2QFire+/gMW/aYDw=",
                    public_key="public_D+CaTZDjs6N/SGKOuhU9pJc4c0M=",
                    url_endpoint="https://ik.imagekit.io/checkmantenimiento/",
                )
                # Foto Entrada - Frente
                if front:
                    file_extension = front.name.split(".")[-1]
                    temp_file = tempfile.NamedTemporaryFile(delete=False)
                    for chunk in front.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    result = imagekit.upload_file(
                        file=open(temp_file.name, "rb"),
                        file_name=f"{equipo.serial}_in_front.{file_extension}",
                        options=UploadFileRequestOptions(folder=f"/{equipo.serial}/"),
                    )
                    equipo.photo_in_front = result.url
                # Foto Entrada - Detras
                if back:
                    file_extension = back.name.split(".")[-1]
                    temp_file = tempfile.NamedTemporaryFile(delete=False)
                    for chunk in back.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    result = imagekit.upload_file(
                        file=open(temp_file.name, "rb"),
                        file_name=f"{equipo.serial}_in_back.{file_extension}",
                        options=UploadFileRequestOptions(folder=f"/{equipo.serial}/"),
                    )
                    equipo.photo_in_back = result.url
                # Foto Entrada - Derecha 1
                if right1:
                    file_extension = right1.name.split(".")[-1]
                    temp_file = tempfile.NamedTemporaryFile(delete=False)
                    for chunk in right1.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    result = imagekit.upload_file(
                        file=open(temp_file.name, "rb"),
                        file_name=f"{equipo.serial}_in_right1.{file_extension}",
                        options=UploadFileRequestOptions(folder=f"/{equipo.serial}/"),
                    )
                    equipo.photo_in_right1 = result.url
                # Foto Entrada - Derecha 2
                if right2:
                    file_extension = right2.name.split(".")[-1]
                    temp_file = tempfile.NamedTemporaryFile(delete=False)
                    for chunk in right2.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    result = imagekit.upload_file(
                        file=open(temp_file.name, "rb"),
                        file_name=f"{equipo.serial}_in_right2.{file_extension}",
                        options=UploadFileRequestOptions(folder=f"/{equipo.serial}/"),
                    )
                    equipo.photo_in_right2 = result.url
                # Foto Entrada - Izquierda 1
                if left1:
                    file_extension = left1.name.split(".")[-1]
                    temp_file = tempfile.NamedTemporaryFile(delete=False)
                    for chunk in left1.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    result = imagekit.upload_file(
                        file=open(temp_file.name, "rb"),
                        file_name=f"{equipo.serial}_in_left1.{file_extension}",
                        options=UploadFileRequestOptions(folder=f"/{equipo.serial}/"),
                    )
                    equipo.photo_in_left1 = result.url
                # Foto Entrada - Izquierda 2
                if left2:
                    file_extension = left2.name.split(".")[-1]
                    temp_file = tempfile.NamedTemporaryFile(delete=False)
                    for chunk in left2.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    result = imagekit.upload_file(
                        file=open(temp_file.name, "rb"),
                        file_name=f"{equipo.serial}_in_left2.{file_extension}",
                        options=UploadFileRequestOptions(folder=f"/{equipo.serial}/"),
                    )
                    equipo.photo_in_left2 = result.url
            equipo.save()
            context["equipo"] = equipo
            context["form"] = form
            if not equipo.photo_in_front:
                error_img = error_img + "frente, "
            if not equipo.photo_in_back:
                error_img = error_img + "detras, "
            if not equipo.photo_in_right1:
                error_img = error_img + "derecha 1, "
            if not equipo.photo_in_right2:
                error_img = error_img + "derecha 2, "
            if not equipo.photo_in_left1:
                error_img = error_img + "izquierda 1, "
            if not equipo.photo_in_left2:
                error_img = error_img + "izquierda 2, "
            if len(error_img) > 29:
                error_img = error_img[:-2]
                context["error_img"] = error_img
            else:
                dibujarequipo(equipo)
                return redirect("software", pk=pk)
    else:
        form = EquipoForm(instance=equipo)
        context["equipo"] = equipo
        context["form"] = form
    return render(request, "mantenimiento/equipo.html", context=context)


@login_required
def software(request, pk):
    equipo = get_object_or_404(Equipo, pk=pk)
    software, created = Software.objects.get_or_create(equipo=equipo)
    if request.method == "POST":
        form = SoftwareForm(request.POST, instance=software)
        if form.is_valid():
            software.save()
            dibujarsoftware(software)
            return redirect("driver", pk=pk)
    else:
        form = SoftwareForm(instance=software)
    return render(request, "mantenimiento/software.html", {"form": form})


@login_required
def driver(request, pk):
    equipo = get_object_or_404(Equipo, pk=pk)
    driver, created = Driver.objects.get_or_create(equipo=equipo)
    if request.method == "POST":
        form = DriverForm(request.POST, instance=driver)
        if form.is_valid():
            driver.save()
            dibujardriver(driver)
            return redirect("revision", pk=pk)
    else:
        form = DriverForm(instance=driver)
    return render(request, "mantenimiento/driver.html", {"form": form})


@login_required
def revision(request, pk):
    equipo = get_object_or_404(Equipo, pk=pk)
    revision, created = Revision.objects.get_or_create(equipo=equipo)
    context = {}
    error_img = "Faltan las siguientes fotos: "
    if request.method == "POST":
        form = RevisionForm(request.POST, instance=revision)
        if form.is_valid():
            revision = form.save(commit=False)
            # Subir imagenes
            front = request.FILES.get("input_out_front")
            back = request.FILES.get("input_out_back")
            right1 = request.FILES.get("input_out_right1")
            right2 = request.FILES.get("input_out_right2")
            left1 = request.FILES.get("input_out_left1")
            left2 = request.FILES.get("input_out_left2")
            if front or back or right1 or right2 or left1 or left2:
                imagekit = ImageKit(
                    private_key="private_OHsLbCwevXZ2QFire+/gMW/aYDw=",
                    public_key="public_D+CaTZDjs6N/SGKOuhU9pJc4c0M=",
                    url_endpoint="https://ik.imagekit.io/checkmantenimiento/",
                )
                # Foto Salida - Frente
                if front:
                    file_extension = front.name.split(".")[-1]
                    temp_file = tempfile.NamedTemporaryFile(delete=False)
                    for chunk in front.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    result = imagekit.upload_file(
                        file=open(temp_file.name, "rb"),
                        file_name=f"{equipo.serial}_out_front.{file_extension}",
                        options=UploadFileRequestOptions(folder=f"/{equipo.serial}/"),
                    )
                    revision.photo_out_front = result.url
                # Foto Salida - Detras
                if back:
                    file_extension = back.name.split(".")[-1]
                    temp_file = tempfile.NamedTemporaryFile(delete=False)
                    for chunk in back.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    result = imagekit.upload_file(
                        file=open(temp_file.name, "rb"),
                        file_name=f"{equipo.serial}_out_back.{file_extension}",
                        options=UploadFileRequestOptions(folder=f"/{equipo.serial}/"),
                    )
                    revision.photo_out_back = result.url
                # Foto Salida - Derecha 1
                if right1:
                    file_extension = right1.name.split(".")[-1]
                    temp_file = tempfile.NamedTemporaryFile(delete=False)
                    for chunk in right1.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    result = imagekit.upload_file(
                        file=open(temp_file.name, "rb"),
                        file_name=f"{equipo.serial}_out_right1.{file_extension}",
                        options=UploadFileRequestOptions(folder=f"/{equipo.serial}/"),
                    )
                    revision.photo_out_right1 = result.url
                # Foto Salida - Derecha 2
                if right2:
                    file_extension = right2.name.split(".")[-1]
                    temp_file = tempfile.NamedTemporaryFile(delete=False)
                    for chunk in right2.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    result = imagekit.upload_file(
                        file=open(temp_file.name, "rb"),
                        file_name=f"{equipo.serial}_out_right2.{file_extension}",
                        options=UploadFileRequestOptions(folder=f"/{equipo.serial}/"),
                    )
                    revision.photo_out_right2 = result.url
                # Foto Salida - Izquierda 1
                if left1:
                    file_extension = left1.name.split(".")[-1]
                    temp_file = tempfile.NamedTemporaryFile(delete=False)
                    for chunk in left1.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    result = imagekit.upload_file(
                        file=open(temp_file.name, "rb"),
                        file_name=f"{equipo.serial}_out_left1.{file_extension}",
                        options=UploadFileRequestOptions(folder=f"/{equipo.serial}/"),
                    )
                    revision.photo_out_left1 = result.url
                # Foto Salida - Izquierda 2
                if left2:
                    file_extension = left2.name.split(".")[-1]
                    temp_file = tempfile.NamedTemporaryFile(delete=False)
                    for chunk in left2.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    result = imagekit.upload_file(
                        file=open(temp_file.name, "rb"),
                        file_name=f"{equipo.serial}_out_left2.{file_extension}",
                        options=UploadFileRequestOptions(folder=f"/{equipo.serial}/"),
                    )
                    revision.photo_out_left2 = result.url
            revision.save()
            context["revision"] = revision
            context["form"] = form
            if not revision.photo_out_front:
                error_img = error_img + "frente, "
            if not revision.photo_out_back:
                error_img = error_img + "detras, "
            if not revision.photo_out_right1:
                error_img = error_img + "derecha 1, "
            if not revision.photo_out_right2:
                error_img = error_img + "derecha 2, "
            if not revision.photo_out_left1:
                error_img = error_img + "izquierda 1, "
            if not revision.photo_out_left2:
                error_img = error_img + "izquierda 2, "
            if len(error_img) > 29:
                error_img = error_img[:-2]
                context["error_img"] = error_img
            else:
                response = dibujarrevision(revision)
                return response
    else:
        form = RevisionForm(instance=revision)
        context["revision"] = revision
        context["form"] = form
    return render(request, "mantenimiento/revision.html", context=context)


def dibujarusuario(usuario):
    current_dir = os.getcwd()
    pdf_path = os.path.join(current_dir, "checkmantenimiento.pythonanywhere.com/checklist.pdf")

    def new_content():
        pdf = FPDF()
        pdf.add_page()

        pdf.set_font("Arial", "", 6)
        pdf.ln(36.5)
        pdf.cell(34)
        y_pos = pdf.get_y()
        pdf.cell(23, 3, usuario.nombre)
        pdf.set_xy(pdf.x + 113.5, y_pos)
        pdf.cell(26.5, 3, usuario.fecha_intervencion.strftime("%d/%m/%y"))

        y_pos = y_pos + 2.8
        pdf.set_xy(67, y_pos)
        pdf.cell(30, 3, usuario.ciudad)
        pdf.set_xy(pdf.x + 65, y_pos)
        pdf.cell(45, 3, usuario.telefono)

        y_pos = y_pos + 2.8
        pdf.set_xy(67, y_pos)
        pdf.cell(30, 3, usuario.usuario_red)
        pdf.set_xy(pdf.x + 65, y_pos)
        pdf.cell(45, 3, usuario.activo)

        return io.BytesIO(pdf.output())

    reader = PdfReader(pdf_path)
    page_overlay = PdfReader(new_content()).pages[0]
    reader.pages[0].merge_page(page2=page_overlay)

    writer = PdfWriter()
    writer.append_pages_from_reader(reader)
    writer.write("checklist-final.pdf")


def dibujarequipo(equipo):
    current_dir = os.getcwd()
    pdf_path = os.path.join(current_dir, "checkmantenimiento.pythonanywhere.com/checklist-final.pdf")

    def new_content():
        pdf = FPDF()
        pdf.add_page()

        pdf.set_font("Arial", "", 6)
        pdf.ln(50.4)
        pdf.cell(34)
        y_pos = pdf.get_y()
        pdf.cell(23, 3, equipo.tipo)
        pdf.set_xy(pdf.x + 15, y_pos)
        pdf.cell(15, 3, equipo.marca)
        pdf.set_xy(pdf.x + 40, y_pos)
        pdf.cell(25, 3, equipo.modelo)
        pdf.set_xy(pdf.x + 19, y_pos)
        pdf.cell(26, 3, equipo.serial)

        pdf.set_font_size(16)
        y_pos = y_pos + 3
        pdf.y = y_pos
        if equipo.estado:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 10, "X", align="C")
        pdf.set_xy(137, y_pos)
        pdf.cell(25.3, 10.5, "6", align="C")

        pdf.set_font_size(6)
        y_pos = y_pos + 9.8
        pdf.set_xy(67, y_pos)
        pdf.cell(30, 3, equipo.sistema_operativo)
        pdf.set_xy(pdf.x + 65, y_pos)
        pdf.cell(45, 3, "Bueno" if equipo.estado_hdd else "Malo")

        y_pos = y_pos + 2.8
        pdf.set_xy(67, y_pos)
        pdf.cell(30, 3, equipo.ram)
        pdf.set_xy(pdf.x + 65, y_pos)
        pdf.cell(45, 3, equipo.tipo_procesador)

        y_pos = y_pos + 5.6
        pdf.set_xy(67, y_pos)
        pdf.multi_cell(70, 2.6, equipo.estado_fisico, align="L")
        pdf.x = pdf.x + 55
        if equipo.siniestro:
            pdf.y = y_pos
        else:
            pdf.y = y_pos + 5.5
        pdf.cell(15, 3, "X")

        return io.BytesIO(pdf.output())

    reader = PdfReader(pdf_path)
    page_overlay = PdfReader(new_content()).pages[0]
    reader.pages[0].merge_page(page2=page_overlay)

    writer = PdfWriter()
    writer.append_pages_from_reader(reader)
    writer.write("checklist-final.pdf")


def dibujarsoftware(software):
    current_dir = os.getcwd()
    pdf_path = os.path.join(current_dir, "checkmantenimiento.pythonanywhere.com/checklist-final.pdf")

    def new_content():
        pdf = FPDF()
        pdf.add_page()

        pdf.set_font("Arial", "", 16)
        pdf.ln(82.6)
        pdf.cell(34)
        y_pos = pdf.get_y()
        if software.des_office:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 10.5, "X", align="C")
        pdf.set_xy(137, y_pos)
        if not software.del_temp2:
            pdf.x = pdf.x + 35
        pdf.cell(25.3, 10.5, "X", align="C")

        y_pos = y_pos + 10.8
        pdf.set_y(y_pos)
        if software.des_chrome:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 10.5, "X", align="C")
        pdf.set_xy(137, y_pos)
        if not software.del_prefetch:
            pdf.x = pdf.x + 35
        pdf.cell(25.3, 10.5, "X", align="C")

        y_pos = y_pos + 10.8
        pdf.set_y(y_pos)
        if software.des_teams:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 10.5, "X", align="C")
        pdf.set_xy(137, y_pos)
        if not software.conf_msconfig:
            pdf.x = pdf.x + 35
        pdf.cell(25.3, 10.5, "X", align="C")

        y_pos = y_pos + 10.8
        pdf.set_y(y_pos)
        if software.del_temp:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 14, "X", align="C")
        pdf.set_xy(137, y_pos)
        if not software.sfc_scannow:
            pdf.x = pdf.x + 35
        pdf.cell(25.3, 14, "X", align="C")

        y_pos = y_pos + 14
        pdf.set_y(y_pos)
        if software.dfg_dd:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 13.5, "X", align="C")
        pdf.set_xy(137, y_pos)
        if not software.win_22h2:
            pdf.x = pdf.x + 35
        pdf.cell(25.3, 13.5, "X", align="C")

        return io.BytesIO(pdf.output())

    reader = PdfReader(pdf_path)
    page_overlay = PdfReader(new_content()).pages[0]
    reader.pages[0].merge_page(page2=page_overlay)

    writer = PdfWriter()
    writer.append_pages_from_reader(reader)
    writer.write("checklist-final.pdf")


def dibujardriver(driver):
    current_dir = os.getcwd()
    pdf_path = os.path.join(current_dir, "checkmantenimiento.pythonanywhere.com/checklist-final.pdf")

    def new_content():
        pdf = FPDF()
        pdf.add_page()

        pdf.set_font("Arial", "", 16)
        pdf.ln(145.3)
        pdf.cell(34)
        y_pos = pdf.get_y()
        if driver.diag_tool:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 8.5, "X", align="C")
        pdf.set_xy(137, y_pos)
        pdf.set_font_size(6)
        pdf.multi_cell(70, 2.6, driver.diag_tool_desc, align="L")

        y_pos = y_pos + 8.5
        pdf.set_y(y_pos)
        if driver.upd_bios:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 2.7, "X", align="C")
        pdf.set_xy(162.3, y_pos)
        pdf.cell(44.8, 2.7, driver.upd_bios_ver)

        y_pos = y_pos + 5.4
        pdf.set_xy(67, y_pos)
        pdf.multi_cell(140, 2.6, driver.upd_bios_desc, align="L")

        pdf.set_font_size(12)
        y_pos = y_pos + 5.7
        pdf.set_y(y_pos)
        if driver.upd_all:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 5.5, "X", align="C")
        pdf.set_font_size(6)
        pdf.set_xy(108, y_pos)
        pdf.multi_cell(99, 2.6, driver.upd_all_desc, align="L")

        y_pos = y_pos + 5.5
        pdf.set_y(y_pos)
        if driver.upd_audio:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 2.7, "====", align="C")
        pdf.set_xy(180.5, y_pos)
        if not driver.upd_almacenamiento:
            pdf.x = pdf.x + 13.3
        pdf.cell(11.8, 2.7, "====", align="C")

        y_pos = y_pos + 2.8
        pdf.set_y(y_pos)
        if driver.upd_video:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 2.7, "====", align="C")
        pdf.set_xy(180.5, y_pos)
        if not driver.upd_chipset:
            pdf.x = pdf.x + 13.3
        pdf.cell(11.8, 2.7, "====", align="C")

        y_pos = y_pos + 2.8
        pdf.set_y(y_pos)
        if driver.upd_wifi:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 2.7, "====", align="C")
        pdf.set_xy(180.5, y_pos)
        if not driver.upd_camara:
            pdf.x = pdf.x + 13.3
        pdf.cell(11.8, 2.7, "====", align="C")

        y_pos = y_pos + 2.9
        pdf.set_y(y_pos)
        if driver.upd_ethernet:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 2.7, "====", align="C")
        pdf.set_xy(180.5, y_pos)
        if not driver.upd_mouse:
            pdf.x = pdf.x + 13.3
        pdf.cell(11.8, 2.7, "====", align="C")

        y_pos = y_pos + 2.9
        pdf.set_y(y_pos)
        if driver.exe_all:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 2.7, "X", align="C")

        y_pos = y_pos + 5.5
        pdf.set_xy(67, y_pos)
        pdf.multi_cell(140, 2.6, driver.observaciones, align="L")

        return io.BytesIO(pdf.output())

    reader = PdfReader(pdf_path)
    page_overlay = PdfReader(new_content()).pages[0]
    reader.pages[0].merge_page(page2=page_overlay)

    writer = PdfWriter()
    writer.append_pages_from_reader(reader)
    writer.write("checklist-final.pdf")


def dibujarrevision(revision):
    current_dir = os.getcwd()
    pdf_path = os.path.join(current_dir, "checkmantenimiento.pythonanywhere.com/checklist-final.pdf")

    def new_content():
        pdf = FPDF()
        pdf.add_page()

        pdf.set_font("Arial", "", 8)
        pdf.ln(198.5)
        pdf.cell(34)
        y_pos = pdf.get_y()
        if revision.teclado:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 2.7, "=====", align="C")
        pdf.set_xy(180.5, y_pos)
        if not revision.wifi:
            pdf.x = pdf.x + 13.3
        pdf.cell(11.8, 2.7, "=====", align="C")

        y_pos = y_pos + 2.8
        pdf.set_y(y_pos)
        if revision.pantalla:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 2.7, "=====", align="C")
        pdf.set_xy(180.5, y_pos)
        if not revision.video_hdmi_vga:
            pdf.x = pdf.x + 13.3
        pdf.cell(11.8, 2.7, "=====", align="C")

        y_pos = y_pos + 2.7
        pdf.set_y(y_pos)
        if revision.sonido:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 5.4, "=====", align="C")
        pdf.set_xy(180.5, y_pos)
        if not revision.internet:
            pdf.x = pdf.x + 13.3
        pdf.cell(11.8, 5.4, "=====", align="C")

        y_pos = y_pos + 5.4
        pdf.set_y(y_pos)
        if revision.camara:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 4, "=====", align="C")
        pdf.set_xy(180.5, y_pos)
        if not revision.apps_cliente:
            pdf.x = pdf.x + 13.3
        pdf.cell(11.8, 4, "=====", align="C")

        y_pos = y_pos + 4
        pdf.set_y(y_pos)
        if revision.touchpad:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 2.7, "=====", align="C")
        pdf.set_xy(180.5, y_pos)
        if not revision.headset:
            pdf.x = pdf.x + 13.3
        pdf.cell(11.8, 2.7, "=====", align="C")

        y_pos = y_pos + 2.7
        pdf.set_y(y_pos)
        if revision.limpieza:
            pdf.x = 67
        else:
            pdf.x = 82
        pdf.cell(15, 2.8, "X", align="C")

        y_pos = y_pos + 2.8
        pdf.set_y(y_pos)
        if revision.funciona:
            pdf.x = 82
        else:
            pdf.x = 109.5
        pdf.cell(15, 6.4, "X", align="C")
        pdf.set_font_size(6)
        pdf.set_xy(147, y_pos)
        pdf.multi_cell(60.2, 2.6, revision.funciona_desc, align="L")

        y_pos = y_pos + 9.7
        pdf.set_xy(16.2, y_pos)
        pdf.multi_cell(191, 2.7, revision.observaciones, align="L")

        pdf.set_font_size(16)
        pdf.set_xy(181.2, 63.2)
        pdf.cell(25.3, 10.5, "6", align="C")

        return io.BytesIO(pdf.output())

    reader = PdfReader(pdf_path)
    page_overlay = PdfReader(new_content()).pages[0]
    reader.pages[0].merge_page(page2=page_overlay)

    writer = PdfWriter()
    writer.append_pages_from_reader(reader)
    # writer.write("checklist-final.pdf")
    pdf_data = io.BytesIO()
    writer.write(pdf_data)
    pdf_data.seek(0)

    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="Checklist-{revision.equipo.serial}.pdf"'
    )

    return response
